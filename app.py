import os
import json
import uuid
import secrets
import shutil
import tempfile
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any
from fastapi import FastAPI, HTTPException, Header, Depends, status
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, DoughnutChart, LineChart, Reference, Series

app = FastAPI(title="Changan Auto Maintenance Dashboard")

DATA_DIR = Path(__file__).parent / "data"
DB_FILE = DATA_DIR / "db.json"
EXAMPLE_DB_FILE = DATA_DIR / "db.example.json"
STATIC_DIR = Path(__file__).parent / "static"

ACTIVE_SESSIONS = set()

def get_admin_password():
    env_pwd = os.environ.get("ADMIN_PASSWORD")
    if env_pwd:
        return env_pwd
    try:
        db = load_db()
        return db.get("admin_password", "admin")
    except Exception:
        return "admin"

def load_db():
    if not DB_FILE.exists():
        if EXAMPLE_DB_FILE.exists():
            shutil.copy2(EXAMPLE_DB_FILE, DB_FILE)
        else:
            raise HTTPException(status_code=500, detail="Database file not found")
    with open(DB_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_db(data):
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def require_admin(authorization: Optional[str] = Header(None)):
    if not authorization:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Требуется авторизация")
    
    token = authorization.replace("Bearer ", "").strip()
    if token not in ACTIVE_SESSIONS:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Сессия истекла или недействительна")
    return True


def get_vehicle_trackers(db, car):
    if "trackers" in car:
        return car["trackers"]
    if car.get("id") == "car_1":
        return db.get("trackers", [])
    return []

def set_vehicle_trackers(db, car, trackers):
    car["trackers"] = trackers
    if car.get("id") == "car_1":
        db["trackers"] = trackers
    # update car inside vehicles array
    vehicles = db.setdefault("vehicles", [])
    for v in vehicles:
        if v.get("id") == car["id"]:
            v["trackers"] = trackers

def get_active_vehicle(db):
    vehicles = db.setdefault("vehicles", [])
    if not vehicles:
        default_car = {
            "id": "car_1",
            "name": "Changan CS55 Plus",
            "brand": "Changan",
            "model": "CS55 Plus",
            "plate": "А 777 АА 777",
            "engine": "1.5T 7DCT",
            "year": 2023,
            "vin": "",
            "current_km": 25340,
            "current_engine_hours": 772,
            "oil_spec": "SAE 0W-20 SP / C5 (4.2 - 4.5 л)"
        }
        vehicles.append(default_car)
        db["active_vehicle_id"] = "car_1"
        save_db(db)
        return default_car
        
    active_id = db.get("active_vehicle_id")
    car = next((v for v in vehicles if v.get("id") == active_id), None)
    if not car:
        car = vehicles[0]
        db["active_vehicle_id"] = car["id"]
        save_db(db)
    return car

class LoginRequest(BaseModel):
    password: str

class PasswordChangeRequest(BaseModel):
    old_password: str
    new_password: str

class VehicleModel(BaseModel):
    id: Optional[str] = None
    brand: str
    model: str
    plate: Optional[str] = ""
    engine: Optional[str] = ""
    year: Optional[int] = None
    vin: Optional[str] = ""
    current_km: int = 0
    current_engine_hours: int = 0
    oil_spec: Optional[str] = ""

class VehicleMileageUpdate(BaseModel):
    current_km: int
    current_engine_hours: int

class MaintenanceRecord(BaseModel):
    vehicle_id: Optional[str] = None
    to_tag: str
    date: str
    engine_hours: int
    mileage: int
    category: str
    item_name: str
    brand: Optional[str] = ""
    article: Optional[str] = ""
    quantity: float = 1.0
    unit: Optional[str] = "шт"
    price_type: Optional[str] = "total"
    price_per_unit: float
    total_price: Optional[float] = None
    interval_km: int = 7500
    interval_hours: int = 250
    note: Optional[str] = "Плановая замена"
    store: Optional[str] = "Ozon"
    url: Optional[str] = ""

class TrackerSetting(BaseModel):
    id: Optional[str] = None
    name: str
    category: str
    match: str
    interval_km: int
    interval_hours: int = 0
    warn_km: int = 1500
    warn_hours: int = 30
    spec: Optional[str] = ""
    article: Optional[str] = ""
    brand: Optional[str] = ""
    icon: Optional[str] = "wrench"
    enabled: bool = True

class PartItem(BaseModel):
    id: Optional[int] = None
    category: str
    item_name: str
    brand: Optional[str] = ""
    article: Optional[str] = ""
    quantity: float = 1.0
    unit: Optional[str] = "шт"
    price_type: Optional[str] = "total"
    price_per_unit: float
    total_price: Optional[float] = None
    interval_km: int = 7500
    interval_hours: int = 250
    note: Optional[str] = "Плановая замена"
    store: Optional[str] = "Ozon"
    url: Optional[str] = ""

class TOGroupPayload(BaseModel):
    vehicle_id: Optional[str] = None
    original_to_tag: Optional[str] = None
    to_tag: str
    date: str
    mileage: int
    engine_hours: int
    parts: List[PartItem]

# --- AUTH ENDPOINTS ---
@app.post("/api/auth/login")
def login(req: LoginRequest):
    expected_pwd = get_admin_password()
    if req.password == expected_pwd:
        token = secrets.token_hex(24)
        ACTIVE_SESSIONS.add(token)
        return {"status": "success", "token": token}
    else:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Неверный пароль")

@app.get("/api/auth/status")
def auth_status(authorization: Optional[str] = Header(None)):
    if authorization:
        token = authorization.replace("Bearer ", "").strip()
        if token in ACTIVE_SESSIONS:
            return {"authenticated": True}
    return {"authenticated": False}

@app.post("/api/auth/logout")
def logout(authorization: Optional[str] = Header(None)):
    if authorization:
        token = authorization.replace("Bearer ", "").strip()
        ACTIVE_SESSIONS.discard(token)
    return {"status": "success"}

@app.post("/api/auth/change-password")
def change_password(req: PasswordChangeRequest, auth: bool = Depends(require_admin)):
    expected_pwd = get_admin_password()
    if req.old_password != expected_pwd:
        raise HTTPException(status_code=400, detail="Текущий пароль неверен")
    if not req.new_password or len(req.new_password) < 3:
        raise HTTPException(status_code=400, detail="Новый пароль должен содержать минимум 3 символа")
    
    db = load_db()
    db["admin_password"] = req.new_password
    save_db(db)
    return {"status": "success", "message": "Пароль успешно изменен"}

# --- VEHICLE / GARAGE ENDPOINTS ---
@app.get("/api/vehicles")
def get_vehicles():
    db = load_db()
    vehicles = db.setdefault("vehicles", [])
    active_car = get_active_vehicle(db)
    return {
        "active_vehicle_id": active_car["id"],
        "active_vehicle": active_car,
        "vehicles": vehicles
    }

@app.post("/api/vehicles/{vehicle_id}/activate")
def activate_vehicle(vehicle_id: str):
    db = load_db()
    vehicles = db.setdefault("vehicles", [])
    car = next((v for v in vehicles if v.get("id") == vehicle_id), None)
    if not car:
        raise HTTPException(status_code=404, detail="Автомобиль не найден")
    db["active_vehicle_id"] = vehicle_id
    save_db(db)
    return {"status": "success", "active_vehicle_id": vehicle_id, "vehicle": car}

@app.post("/api/vehicles")
def add_vehicle(v: VehicleModel, auth: bool = Depends(require_admin)):
    db = load_db()
    vehicles = db.setdefault("vehicles", [])
    new_id = f"car_{len(vehicles) + 1}_{secrets.token_hex(2)}"
    
    new_car = {
        "id": new_id,
        "name": f"{v.brand} {v.model}",
        "brand": v.brand,
        "model": v.model,
        "plate": v.plate or "",
        "engine": v.engine or "",
        "year": v.year,
        "vin": v.vin or "",
        "current_km": v.current_km or 0,
        "current_engine_hours": v.current_engine_hours or 0,
        "oil_spec": v.oil_spec or "",
        "trackers": []
    }
    vehicles.append(new_car)
    db["active_vehicle_id"] = new_id
    save_db(db)
    return {"status": "success", "vehicle": new_car}

@app.put("/api/vehicles/{vehicle_id}")
def update_vehicle(vehicle_id: str, v: VehicleModel, auth: bool = Depends(require_admin)):
    db = load_db()
    vehicles = db.setdefault("vehicles", [])
    idx = next((i for i, car in enumerate(vehicles) if car.get("id") == vehicle_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="Автомобиль не найден")
    
    vehicles[idx]["brand"] = v.brand
    vehicles[idx]["model"] = v.model
    vehicles[idx]["name"] = f"{v.brand} {v.model}"
    vehicles[idx]["plate"] = v.plate or ""
    vehicles[idx]["engine"] = v.engine or ""
    vehicles[idx]["year"] = v.year
    vehicles[idx]["vin"] = v.vin or ""
    vehicles[idx]["current_km"] = v.current_km
    vehicles[idx]["current_engine_hours"] = v.current_engine_hours
    vehicles[idx]["oil_spec"] = v.oil_spec or ""
    
    save_db(db)
    return {"status": "success", "vehicle": vehicles[idx]}

@app.delete("/api/vehicles/{vehicle_id}")
def delete_vehicle(vehicle_id: str, auth: bool = Depends(require_admin)):
    db = load_db()
    vehicles = db.setdefault("vehicles", [])
    if len(vehicles) <= 1:
        raise HTTPException(status_code=400, detail="Нельзя удалить единственный автомобиль в гараже")
        
    db["vehicles"] = [v for v in vehicles if v.get("id") != vehicle_id]
    db["maintenance_records"] = [r for r in db.get("maintenance_records", []) if r.get("vehicle_id") != vehicle_id]
    
    if db.get("active_vehicle_id") == vehicle_id:
        db["active_vehicle_id"] = db["vehicles"][0]["id"]
        
    save_db(db)
    return {"status": "deleted", "active_vehicle_id": db["active_vehicle_id"]}

@app.post("/api/vehicle/mileage")
@app.post("/api/vehicle")
def update_mileage(v: VehicleMileageUpdate, auth: bool = Depends(require_admin)):
    db = load_db()
    car = get_active_vehicle(db)
    car["current_km"] = v.current_km
    car["current_engine_hours"] = v.current_engine_hours
    
    # Also update any vehicles list and legacy vehicle object
    if "vehicles" in db:
        for item in db["vehicles"]:
            if item.get("id") == car["id"]:
                item["current_km"] = v.current_km
                item["current_engine_hours"] = v.current_engine_hours
    db["vehicle"] = car
    save_db(db)
    return {"status": "success", "vehicle": car}


# --- STATUS & RECORDS ENDPOINTS (SCOPED BY ACTIVE VEHICLE) ---
@app.get("/")
def get_index():
    return FileResponse(STATIC_DIR / "index.html")

@app.get("/api/status")
def get_status():
    db = load_db()
    vehicle = get_active_vehicle(db)
    v_id = vehicle["id"]
    
    all_records = db.get("maintenance_records", [])
    records = [r for r in all_records if r.get("vehicle_id", "car_1") == v_id]
    trackers = get_vehicle_trackers(db, vehicle)
    
    current_km = vehicle.get("current_km", 0)
    current_hours = vehicle.get("current_engine_hours", 0)
    
    total_spent = sum(r.get("total_price", 0) for r in records)
    cost_per_km = round(total_spent / current_km, 2) if current_km > 0 else 0
    avg_speed = round(current_km / current_hours, 1) if current_hours > 0 else 0
    
    # Expense Breakdown
    to_spent = sum(r.get("total_price", 0) for r in records if str(r.get("to_tag", "")).upper().startswith("ТО"))
    custom_spent = total_spent - to_spent
    tuning_spent = sum(r.get("total_price", 0) for r in records if "тюнинг" in str(r.get("category", "")).lower() or "тюнинг" in str(r.get("to_tag", "")).lower())
    tires_spent = sum(r.get("total_price", 0) for r in records if "шин" in str(r.get("category", "")).lower() or "колес" in str(r.get("category", "")).lower())
    
    consumables_status = []
    
    for tracker in trackers:
        if not tracker.get("enabled", True):
            continue
            
        keyword = tracker.get("match", "").lower()
        matching = [r for r in records if keyword in r["item_name"].lower()]
        matching.sort(key=lambda x: (x["mileage"], x["date"]))
        latest = matching[-1] if matching else None
        
        interval_km = tracker.get("interval_km", 7500)
        interval_h = tracker.get("interval_hours", 250)
        warn_km = tracker.get("warn_km", 1500)
        warn_h = tracker.get("warn_hours", 30)
        
        if latest:
            last_km = latest["mileage"]
            last_h = latest["engine_hours"]
            rec_interval_km = latest.get("interval_km") or interval_km
            rec_interval_h = latest.get("interval_hours") or interval_h
            
            next_km = latest.get("next_km", last_km + rec_interval_km)
            next_h = latest.get("next_hours", last_h + rec_interval_h if rec_interval_h > 0 else 0)
            
            # Effective current mileage for this consumable cannot be less than replacement point
            eff_current_km = max(current_km, last_km)
            eff_current_h = max(current_hours, last_h)
            
            rem_km = next_km - eff_current_km
            rem_h = (next_h - eff_current_h) if next_h > 0 else None
            
            used_km = eff_current_km - last_km
            wear_ratio = max(0.0, min(1.0, used_km / rec_interval_km)) if rec_interval_km > 0 else 0.0
            wear_percent = round(wear_ratio * 100, 1)
            
            if rem_km <= 0 or (rem_h is not None and rem_h <= 0):
                status_code = "danger"
                status_text = "Требуется замена"
            elif rem_km <= warn_km or (rem_h is not None and rem_h <= warn_h):
                status_code = "warning"
                status_text = "Скоро замена"
            else:
                status_code = "ok"
                status_text = "В норме"
                
            consumables_status.append({
                "tracker_id": tracker.get("id"),
                "item_name": tracker.get("name"),
                "icon": tracker.get("icon", "wrench"),
                "category": tracker.get("category", "Прочее"),
                "last_date": latest["date"],
                "last_km": last_km,
                "last_hours": last_h,
                "next_km": next_km,
                "next_hours": next_h,
                "rem_km": rem_km,
                "rem_hours": rem_h,
                "wear_percent": wear_percent,
                "status_code": status_code,
                "status_text": status_text,
                "brand": latest.get("brand") or tracker.get("brand", ""),
                "article": latest.get("article") or tracker.get("article", ""),
                "spec": tracker.get("spec", ""),
                "to_tag": latest.get("to_tag", "")
            })
        else:
            consumables_status.append({
                "tracker_id": tracker.get("id"),
                "item_name": tracker.get("name"),
                "icon": tracker.get("icon", "wrench"),
                "category": tracker.get("category", "Прочее"),
                "last_date": "-",
                "last_km": 0,
                "last_hours": 0,
                "next_km": interval_km,
                "next_hours": interval_h,
                "rem_km": interval_km - current_km,
                "rem_hours": (interval_h - current_hours) if interval_h > 0 else None,
                "wear_percent": 0.0,
                "status_code": "warning" if current_km >= interval_km else "ok",
                "status_text": "Нет записей ТО",
                "brand": tracker.get("brand", ""),
                "article": tracker.get("article", ""),
                "spec": tracker.get("spec", ""),
                "to_tag": "-"
            })
    
    attention_count = sum(1 for c in consumables_status if c["status_code"] in ["danger", "warning"])
    
    return {
        "vehicle": vehicle,
        "kpi": {
            "current_km": current_km,
            "current_hours": current_hours,
            "total_spent": total_spent,
            "cost_per_km": cost_per_km,
            "avg_speed": avg_speed,
            "total_records": len(records),
            "attention_count": attention_count
        },
        "consumables": consumables_status
    }

@app.get("/api/records")
def get_records():
    db = load_db()
    v_id = get_active_vehicle(db)["id"]
    return [r for r in db.get("maintenance_records", []) if r.get("vehicle_id", "car_1") == v_id]

@app.get("/api/to-groups")
def get_to_groups():
    db = load_db()
    v_id = get_active_vehicle(db)["id"]
    records = [r for r in db.get("maintenance_records", []) if r.get("vehicle_id", "car_1") == v_id]
    
    groups: Dict[str, Any] = {}
    for r in records:
        tag = r.get("to_tag") or "Без метки"
        if tag not in groups:
            groups[tag] = {
                "to_tag": tag,
                "date": r["date"],
                "mileage": r["mileage"],
                "engine_hours": r["engine_hours"],
                "total_cost": 0,
                "parts_count": 0,
                "parts": []
            }
        groups[tag]["total_cost"] += r.get("total_price", 0)
        groups[tag]["parts_count"] += 1
        groups[tag]["parts"].append(r)
        
    return list(groups.values())

@app.get("/api/settings")
def get_settings():
    db = load_db()
    vehicle = get_active_vehicle(db)
    return {
        "vehicle": vehicle,
        "trackers": get_vehicle_trackers(db, vehicle)
    }

# --- WRITE ENDPOINTS (PROTECTED) ---
@app.post("/api/records")
def add_record(record: MaintenanceRecord, auth: bool = Depends(require_admin)):
    db = load_db()
    car = get_active_vehicle(db)
    v_id = car["id"]
    
    records = db.setdefault("maintenance_records", [])
    new_id = (max([r["id"] for r in records]) + 1) if records else 1
    
    total_price = record.total_price if record.total_price is not None else record.price_per_unit
    next_km = record.mileage + record.interval_km
    next_hours = (record.engine_hours + record.interval_hours) if record.interval_hours > 0 else 0
    
    new_rec = {
        "id": new_id,
        "vehicle_id": v_id,
        "to_tag": record.to_tag,
        "date": record.date,
        "engine_hours": record.engine_hours,
        "mileage": record.mileage,
        "category": record.category,
        "item_name": record.item_name,
        "brand": record.brand or "",
        "article": record.article or "",
        "quantity": record.quantity,
        "unit": record.unit or "шт",
        "price_type": record.price_type or "total",
        "price_per_unit": record.price_per_unit,
        "total_price": total_price,
        "interval_km": record.interval_km,
        "interval_hours": record.interval_hours,
        "next_km": next_km,
        "next_hours": next_hours,
        "note": record.note or "",
        "store": record.store or "",
        "url": record.url or ""
    }
    
    records.append(new_rec)
    
    if record.mileage > car.get("current_km", 0):
        car["current_km"] = record.mileage
    if record.engine_hours > car.get("current_engine_hours", 0):
        car["current_engine_hours"] = record.engine_hours
        
    save_db(db)
    return {"status": "success", "record": new_rec}

@app.put("/api/records/{record_id}")
def update_record(record_id: int, record: MaintenanceRecord, auth: bool = Depends(require_admin)):
    db = load_db()
    car = get_active_vehicle(db)
    records = db.get("maintenance_records", [])
    idx = next((i for i, r in enumerate(records) if r["id"] == record_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    
    total_price = record.total_price if record.total_price is not None else record.price_per_unit
    next_km = record.mileage + record.interval_km
    next_hours = (record.engine_hours + record.interval_hours) if record.interval_hours > 0 else 0
    
    updated_rec = {
        "id": record_id,
        "vehicle_id": records[idx].get("vehicle_id", car["id"]),
        "to_tag": record.to_tag,
        "date": record.date,
        "engine_hours": record.engine_hours,
        "mileage": record.mileage,
        "category": record.category,
        "item_name": record.item_name,
        "brand": record.brand or "",
        "article": record.article or "",
        "quantity": record.quantity,
        "unit": record.unit or "шт",
        "price_type": record.price_type or "total",
        "price_per_unit": record.price_per_unit,
        "total_price": total_price,
        "interval_km": record.interval_km,
        "interval_hours": record.interval_hours,
        "next_km": next_km,
        "next_hours": next_hours,
        "note": record.note or "",
        "store": record.store or "",
        "url": record.url or ""
    }
    
    records[idx] = updated_rec
    save_db(db)
    return {"status": "success", "record": updated_rec}

@app.delete("/api/records/{record_id}")
def delete_record(record_id: int, auth: bool = Depends(require_admin)):
    db = load_db()
    records = db.get("maintenance_records", [])
    db["maintenance_records"] = [r for r in records if r["id"] != record_id]
    save_db(db)
    return {"status": "deleted", "id": record_id}

@app.post("/api/to-groups")
def save_to_group(payload: TOGroupPayload, auth: bool = Depends(require_admin)):
    db = load_db()
    car = get_active_vehicle(db)
    v_id = car["id"]
    
    records = db.setdefault("maintenance_records", [])
    
    if payload.original_to_tag:
        db["maintenance_records"] = [r for r in records if not (r.get("vehicle_id", "car_1") == v_id and r.get("to_tag") == payload.original_to_tag)]
        records = db["maintenance_records"]
        
    current_max_id = max([r["id"] for r in records]) if records else 0
    
    for part in payload.parts:
        current_max_id += 1
        total_p = part.total_price if part.total_price is not None else part.price_per_unit
        next_k = payload.mileage + part.interval_km
        next_h = (payload.engine_hours + part.interval_hours) if part.interval_hours > 0 else 0
        
        new_item = {
            "id": current_max_id,
            "vehicle_id": v_id,
            "to_tag": payload.to_tag,
            "date": payload.date,
            "engine_hours": payload.engine_hours,
            "mileage": payload.mileage,
            "category": part.category,
            "item_name": part.item_name,
            "brand": part.brand or "",
            "article": part.article or "",
            "quantity": part.quantity,
            "unit": part.unit or "шт",
            "price_type": part.price_type or "total",
            "price_per_unit": part.price_per_unit,
            "total_price": total_p,
            "interval_km": part.interval_km,
            "interval_hours": part.interval_hours,
            "next_km": next_k,
            "next_hours": next_h,
            "note": part.note or "",
            "store": part.store or "",
            "url": part.url or ""
        }
        records.append(new_item)
        
    if payload.mileage > car.get("current_km", 0):
        car["current_km"] = payload.mileage
    if payload.engine_hours > car.get("current_engine_hours", 0):
        car["current_engine_hours"] = payload.engine_hours
        
    save_db(db)
    return {"status": "success", "to_tag": payload.to_tag, "parts_added": len(payload.parts)}

@app.delete("/api/to-groups/{to_tag}")
def delete_to_group(to_tag: str, auth: bool = Depends(require_admin)):
    db = load_db()
    v_id = get_active_vehicle(db)["id"]
    records = db.get("maintenance_records", [])
    db["maintenance_records"] = [r for r in records if not (r.get("vehicle_id", "car_1") == v_id and r.get("to_tag") == to_tag)]
    save_db(db)
    return {"status": "deleted", "to_tag": to_tag}

@app.post("/api/settings/tracker")
def save_tracker(tracker: TrackerSetting, auth: bool = Depends(require_admin)):
    db = load_db()
    car = get_active_vehicle(db)
    trackers = list(get_vehicle_trackers(db, car))
    
    t_id = tracker.id or tracker.name.lower().replace(" ", "_").replace("(", "").replace(")", "")
    idx = next((i for i, t in enumerate(trackers) if t.get("id") == t_id), None)
    
    tracker_dict = {
        "id": t_id,
        "name": tracker.name,
        "category": tracker.category,
        "match": tracker.match,
        "interval_km": tracker.interval_km,
        "interval_hours": tracker.interval_hours,
        "warn_km": tracker.warn_km,
        "warn_hours": tracker.warn_hours,
        "spec": tracker.spec or "",
        "article": tracker.article or "",
        "brand": tracker.brand or "",
        "icon": tracker.icon or "wrench",
        "enabled": tracker.enabled
    }
    
    if idx is not None:
        trackers[idx] = tracker_dict
    else:
        trackers.append(tracker_dict)
        
    set_vehicle_trackers(db, car, trackers)
    save_db(db)
    return {"status": "success", "tracker": tracker_dict}

@app.delete("/api/settings/tracker/{tracker_id}")
def delete_tracker(tracker_id: str, auth: bool = Depends(require_admin)):
    db = load_db()
    trackers = get_vehicle_trackers(db, vehicle)
    db["trackers"] = [t for t in trackers if t.get("id") != tracker_id]
    save_db(db)
    return {"status": "deleted", "id": tracker_id}

@app.get("/api/export-excel")
def export_excel():
    db = load_db()
    vehicle = get_active_vehicle(db)
    v_id = vehicle["id"]
    all_records = db.get("maintenance_records", [])
    records = [r for r in all_records if r.get("vehicle_id", "car_1") == v_id]
    trackers = get_vehicle_trackers(db, vehicle)
    
    current_km = vehicle.get("current_km", 0)
    current_hours = vehicle.get("current_engine_hours", 0)
    total_spent = sum(r.get("total_price", 0) for r in records)
    cost_per_km = round(total_spent / current_km, 2) if current_km > 0 else 0
    avg_speed = round(current_km / current_hours, 1) if current_hours > 0 else 0
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    NAVY_HEADER = "1E293B"
    BLUE_ACCENT = "2563EB"
    INDIGO_HEADER = "4338CA"
    EMERALD_COLOR = "059669"
    WHITE = "FFFFFF"
    BORDER_COLOR = "CBD5E1"
    LIGHT_BG = "F8FAFC"
    CARD_BG = "F1F5F9"
    
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    thick_bottom_navy = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='medium', color=NAVY_HEADER)
    )
    
    font_title = Font(name="Segoe UI", size=14, bold=True, color="0F172A")
    font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    font_section = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
    font_header = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_regular = Font(name="Segoe UI", size=10)
    font_kpi_value = Font(name="Segoe UI", size=13, bold=True, color="1E293B")
    font_kpi_label = Font(name="Segoe UI", size=8, bold=True, color="64748B")
    
    fill_header_navy = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    fill_header_indigo = PatternFill(start_color=INDIGO_HEADER, end_color=INDIGO_HEADER, fill_type="solid")
    fill_header_blue = PatternFill(start_color=BLUE_ACCENT, end_color=BLUE_ACCENT, fill_type="solid")
    fill_card = PatternFill(start_color=CARD_BG, end_color=CARD_BG, fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    # -------------------------------------------------------------
    # SHEET 1: АНАЛИТИКА И ГРАФИКИ
    # -------------------------------------------------------------
    ws_charts = wb.create_sheet(title="Аналитика и Графики")
    ws_charts.views.sheetView[0].showGridLines = True
    
    # Title Header Block
    ws_charts["B2"] = f"ОТЧЕТ ПО ОБСЛУЖИВАНИЮ: {vehicle.get('brand', '')} {vehicle.get('model', '')}"
    ws_charts["B2"].font = font_title
    ws_charts["B3"] = f"Госномер: {vehicle.get('plate', '-')} | Двигатель: {vehicle.get('engine', '-')} | Автор проекта: Щеголев Александр (scanek)"
    ws_charts["B3"].font = font_subtitle
    
    # KPI Mini Cards (Row 5-6)
    kpis = [
        ("ТЕКУЩИЙ ПРОБЕГ", f"{current_km:,} км".replace(",", " "), "B"),
        ("МОТОЧАСЫ", f"{current_hours:,} м/ч".replace(",", " "), "C"),
        ("ВСЕГО ЗАТРАТ", f"{total_spent:,} ₽".replace(",", " "), "D"),
        ("СТОИМОСТЬ КМ", f"{cost_per_km} ₽/км", "E"),
        ("СРЕДНЯЯ СКОРОСТЬ", f"{avg_speed} км/ч", "F"),
    ]
    for label, val, col in kpis:
        c_lbl = ws_charts[f"{col}5"]
        c_lbl.value = label
        c_lbl.font = font_kpi_label
        c_lbl.alignment = Alignment(horizontal="center", vertical="center")
        c_lbl.fill = fill_card
        c_lbl.border = thin_border
        
        c_val = ws_charts[f"{col}6"]
        c_val.value = val
        c_val.font = font_kpi_value
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.fill = fill_card
        c_val.border = thin_border
        
    # --- TABLE 1: Категории расходов (Data Source for Pie/Doughnut Chart) ---
    ws_charts["B9"] = "Структура расходов по категориям"
    ws_charts["B9"].font = font_section
    
    cat_headers = ["Категория", "Сумма (₽)", "Доля"]
    for ci, h in enumerate(cat_headers, start=2):
        cell = ws_charts.cell(row=10, column=ci, value=h)
        cell.font = font_header
        cell.fill = fill_header_indigo
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    cat_totals = {}
    for r in records:
        cat = r.get("category") or "Прочее"
        tag = str(r.get("to_tag", ""))
        if "тюнинг" in tag.lower() or "тюнинг" in cat.lower():
            cat = "Тюнинг и доработки"
        elif "шин" in tag.lower() or "шин" in cat.lower() or "колес" in cat.lower():
            cat = "Шины и колеса"
        elif "акб" in tag.lower() or "аккумулятор" in cat.lower():
            cat = "Аккумулятор и электрика"
        elif "ремонт" in tag.lower() or "ремонт" in cat.lower():
            cat = "Ремонт и сервис"
        cat_totals[cat] = cat_totals.get(cat, 0) + r.get("total_price", 0)
        
    if not cat_totals:
        cat_totals["Плановое ТО"] = 0
        
    cat_row_start = 11
    cat_row_end = cat_row_start + len(cat_totals) - 1
    
    for idx, (cat_name, sum_val) in enumerate(sorted(cat_totals.items(), key=lambda x: x[1], reverse=True), start=cat_row_start):
        c1 = ws_charts.cell(row=idx, column=2, value=cat_name)
        c2 = ws_charts.cell(row=idx, column=3, value=sum_val)
        c3 = ws_charts.cell(row=idx, column=4, value=f"=C{idx}/C{cat_row_end+1}" if total_spent > 0 else 0)
        
        c1.font = font_regular
        c2.font = font_bold
        c3.font = font_regular
        
        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border
        
        c2.number_format = '#,##0 "₽"'
        c3.number_format = '0.0%'
        c3.alignment = Alignment(horizontal="center")
        
    # Total row for categories
    tot_row = cat_row_end + 1
    ws_charts.cell(row=tot_row, column=2, value="ИТОГО:").font = font_bold
    ws_charts.cell(row=tot_row, column=2).fill = fill_total
    ws_charts.cell(row=tot_row, column=2).border = thin_border
    
    c_tot_val = ws_charts.cell(row=tot_row, column=3, value=f"=SUM(C{cat_row_start}:C{cat_row_end})")
    c_tot_val.font = font_bold
    c_tot_val.fill = fill_total
    c_tot_val.border = thin_border
    c_tot_val.number_format = '#,##0 "₽"'
    
    c_tot_pct = ws_charts.cell(row=tot_row, column=4, value=1.0)
    c_tot_pct.font = font_bold
    c_tot_pct.fill = fill_total
    c_tot_pct.border = thin_border
    c_tot_pct.number_format = '0.0%'
    c_tot_pct.alignment = Alignment(horizontal="center")
    
    # --- CHART 1: Doughnut Chart (Категории расходов) ---
    pie = DoughnutChart()
    pie.title = "Структура расходов автомобиля"
    pie.style = 10
    labels = Reference(ws_charts, min_col=2, min_row=cat_row_start, max_row=cat_row_end)
    data = Reference(ws_charts, min_col=3, min_row=10, max_row=cat_row_end)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 15
    pie.height = 8.5
    pie.holeSize = 40
    ws_charts.add_chart(pie, "F9")
    
    # --- TABLE 2: Динамика расходов по событиям (Data Source for Column BarChart) ---
    event_start_row = max(tot_row + 3, 20)
    ws_charts.cell(row=event_start_row - 1, column=2, value="Расходы по событиям ТО и покупкам").font = font_section
    
    ev_headers = ["Событие / ТО", "Пробег (км)", "Затраты (₽)"]
    for ci, h in enumerate(ev_headers, start=2):
        cell = ws_charts.cell(row=event_start_row, column=ci, value=h)
        cell.font = font_header
        cell.fill = fill_header_blue
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    # Aggregate by event tag
    events_map = {}
    for r in records:
        tag = r.get("to_tag") or "Вне ТО"
        if tag not in events_map:
            events_map[tag] = {"tag": tag, "mileage": r.get("mileage", 0), "total": 0}
        events_map[tag]["total"] += r.get("total_price", 0)
        events_map[tag]["mileage"] = max(events_map[tag]["mileage"], r.get("mileage", 0))
        
    ev_data_start = event_start_row + 1
    ev_data_end = ev_data_start + max(len(events_map), 1) - 1
    
    if events_map:
        for idx, ev in enumerate(events_map.values(), start=ev_data_start):
            c1 = ws_charts.cell(row=idx, column=2, value=ev["tag"])
            c2 = ws_charts.cell(row=idx, column=3, value=ev["mileage"])
            c3 = ws_charts.cell(row=idx, column=4, value=ev["total"])
            
            c1.font = font_bold
            c2.font = font_regular
            c3.font = font_bold
            
            c1.border = thin_border
            c2.border = thin_border
            c3.border = thin_border
            
            c2.number_format = '#,##0 "км"'
            c3.number_format = '#,##0 "₽"'
            c2.alignment = Alignment(horizontal="center")
    else:
        ws_charts.cell(row=ev_data_start, column=2, value="Нет записей").border = thin_border
        ws_charts.cell(row=ev_data_start, column=3, value=0).border = thin_border
        ws_charts.cell(row=ev_data_start, column=4, value=0).border = thin_border
        
    # --- CHART 2: Column Chart (Динамика расходов) ---
    bar = BarChart()
    bar.type = "col"
    bar.style = 11
    bar.title = "Затраты по событиям обслуживания (₽)"
    bar.y_axis.title = "Сумма (₽)"
    bar.x_axis.title = "Событие"
    data_bar = Reference(ws_charts, min_col=4, min_row=event_start_row, max_row=ev_data_end)
    labels_bar = Reference(ws_charts, min_col=2, min_row=ev_data_start, max_row=ev_data_end)
    bar.add_data(data_bar, titles_from_data=True)
    bar.set_categories(labels_bar)
    bar.legend = None # No legend needed for single series
    bar.width = 15
    bar.height = 8.5
    ws_charts.add_chart(bar, f"F{event_start_row}")
    
    # -------------------------------------------------------------
    # SHEET 2: СОСТОЯНИЕ РАСХОДНИКОВ И РЕСУРС
    # -------------------------------------------------------------
    ws_status = wb.create_sheet(title="Состояние расходников")
    ws_status.views.sheetView[0].showGridLines = True
    
    ws_status["B2"] = f"ТЕКУЩИЙ ИЗНОС И РЕСУРС РАСХОДНЫХ МАТЕРИАЛОВ: {vehicle.get('name', '')}"
    ws_status["B2"].font = font_title
    ws_status["B3"] = f"Текущий пробег: {current_km:,} км | Моточасы: {current_hours:,} м/ч".replace(",", " ")
    ws_status["B3"].font = font_subtitle
    
    st_headers = [
        "Расходник", "Категория", "Последняя замена", "Пробег зам. (км)", "Моточасы (м/ч)",
        "След. замена (км)", "След. замена (м/ч)", "Остаток (км)", "Остаток (м/ч)",
        "Износ (%)", "Статус", "Артикул / Спецификация"
    ]
    for ci, h in enumerate(st_headers, start=2):
        cell = ws_status.cell(row=5, column=ci, value=h)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_bottom_navy
        
    for r_idx, tracker in enumerate(trackers, start=6):
        keyword = tracker.get("match", "").lower()
        matching = [r for r in records if keyword in r["item_name"].lower()]
        matching.sort(key=lambda x: (x["mileage"], x["date"]))
        latest = matching[-1] if matching else None
        
        interval_km = tracker.get("interval_km", 7500)
        interval_h = tracker.get("interval_hours", 250)
        
        if latest:
            last_date = latest["date"]
            last_km = latest["mileage"]
            last_h = latest["engine_hours"]
            rec_int_km = latest.get("interval_km") or interval_km
            rec_int_h = latest.get("interval_hours") or interval_h
            
            next_km = latest.get("next_km", last_km + rec_int_km)
            next_h = latest.get("next_hours", last_h + rec_int_h if rec_int_h > 0 else 0)
            
            eff_km = max(current_km, last_km)
            eff_h = max(current_hours, last_h)
            rem_km = next_km - eff_km
            rem_h = (next_h - eff_h) if next_h > 0 else None
            
            used_km = eff_km - last_km
            wear = max(0.0, min(1.0, used_km / rec_int_km)) if rec_int_km > 0 else 0.0
            
            if rem_km <= 0 or (rem_h is not None and rem_h <= 0):
                st_code = "ТРЕБУЕТСЯ ЗАМЕНА!"
                st_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                st_font = Font(name="Segoe UI", size=9, bold=True, color="991B1B")
            elif rem_km <= tracker.get("warn_km", 1500) or (rem_h is not None and rem_h <= tracker.get("warn_hours", 30)):
                st_code = "Внимание (скоро замена)"
                st_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                st_font = Font(name="Segoe UI", size=9, bold=True, color="92400E")
            else:
                st_code = "В норме"
                st_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                st_font = Font(name="Segoe UI", size=9, bold=True, color="065F46")
        else:
            last_date = "Нет данных"
            last_km = 0
            last_h = 0
            next_km = interval_km
            next_h = interval_h
            rem_km = interval_km - current_km
            rem_h = (interval_h - current_hours) if interval_h > 0 else None
            wear = 0.0
            st_code = "Нет записей ТО"
            st_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            st_font = Font(name="Segoe UI", size=9, bold=True, color="475569")
            
        row_vals = [
            tracker.get("name", ""),
            tracker.get("category", "Прочее"),
            last_date,
            last_km,
            last_h if last_h > 0 else "-",
            next_km,
            next_h if next_h > 0 else "-",
            rem_km,
            rem_h if rem_h is not None else "-",
            wear,
            st_code,
            tracker.get("article") or tracker.get("spec") or "-"
        ]
        
        for ci, val in enumerate(row_vals, start=2):
            cell = ws_status.cell(row=r_idx, column=ci, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if ci in [4, 6, 8]:
                cell.number_format = '#,##0 "км"'
                cell.alignment = Alignment(horizontal="right")
            elif ci == 11:
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="center")
            elif ci == 12:
                cell.fill = st_fill
                cell.font = st_font
                cell.alignment = Alignment(horizontal="center")
            elif ci in [2, 3, 5, 7, 9, 10]:
                cell.alignment = Alignment(horizontal="center")
                
    # -------------------------------------------------------------
    # SHEET 3: ЖУРНАЛ ВСЕХ РАСХОДОВ И ТО
    # -------------------------------------------------------------
    ws_log = wb.create_sheet(title="Журнал всех записей")
    ws_log.views.sheetView[0].showGridLines = True
    
    ws_log["B2"] = f"ПОЛНЫЙ ЖУРНАЛ ТО, ТЮНИНГА И ПОКУПОК: {vehicle.get('name', '')}"
    ws_log["B2"].font = font_title
    ws_log["B3"] = f"Всего позиций: {len(records)} | Общая сумма затрат: {total_spent:,} ₽".replace(",", " ")
    ws_log["B3"].font = font_subtitle
    
    log_headers = [
        "Метка / ТО", "№", "Дата", "Пробег (км)", "Моточасы (м/ч)", "Категория",
        "Наименование детали / работы", "Бренд", "Артикул / Модель", "Кол-во", "Ед.",
        "Цена за ед.", "Сумма (₽)", "След. замена (км)", "След. замена (м/ч)",
        "Магазин / Сервис", "Заметки / Гарантия", "Ссылка"
    ]
    for ci, h in enumerate(log_headers, start=2):
        cell = ws_log.cell(row=5, column=ci, value=h)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_bottom_navy
        
    for r_idx, r in enumerate(records, start=6):
        is_zebra = (r_idx % 2 == 0)
        row_fill = fill_zebra if is_zebra else None
        
        tag = r.get("to_tag") or "Вне ТО"
        
        row_vals = [
            tag,
            r_idx - 5,
            r.get("date", ""),
            r.get("mileage", 0),
            r.get("engine_hours", 0) if r.get("engine_hours", 0) > 0 else "-",
            r.get("category", "Прочее"),
            r.get("item_name", ""),
            r.get("brand", "") or "-",
            r.get("article", "") or "-",
            r.get("quantity", 1),
            r.get("unit", "шт"),
            r.get("price_per_unit", 0),
            r.get("total_price", 0),
            r.get("next_km", 0) if r.get("next_km", 0) > 0 else "-",
            r.get("next_hours", 0) if r.get("next_hours", 0) > 0 else "-",
            r.get("store", "") or "-",
            r.get("note", "") or "-",
            r.get("url", "") or "-"
        ]
        
        for ci, val in enumerate(row_vals, start=2):
            cell = ws_log.cell(row=r_idx, column=ci, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
                
            if ci in [5, 15]:
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0 "км"'
                cell.alignment = Alignment(horizontal="right")
            elif ci in [13, 14]:
                cell.number_format = '#,##0 "₽"'
                cell.font = font_bold
                cell.alignment = Alignment(horizontal="right")
            elif ci in [2, 3, 4, 6, 11, 12]:
                cell.alignment = Alignment(horizontal="center")
                if ci == 2:
                    cell.font = font_bold
                    
    # Log Total row
    log_tot_row = len(records) + 6
    c_itogo = ws_log.cell(row=log_tot_row, column=2, value="ИТОГО:")
    c_itogo.font = font_bold
    c_itogo.fill = fill_total
    c_itogo.border = thin_border
    c_itogo.alignment = Alignment(horizontal="center", vertical="center")
    
    for ci in range(3, 20):
        c = ws_log.cell(row=log_tot_row, column=ci)
        c.fill = fill_total
        c.border = thin_border
        if ci == 14: # Column N (Сумма ₽)
            c.value = f"=SUM(N6:N{log_tot_row-1})"
            c.font = Font(name="Segoe UI", size=10, bold=True, color="047857")
            c.number_format = '#,##0 "₽"'
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif ci == 3: # Column C (Количество записей)
            c.value = f"{len(records)} поз."
            c.font = font_bold
            c.alignment = Alignment(horizontal="center", vertical="center")
            
    # -------------------------------------------------------------
    # SHEET 4: РЕГЛАМЕНТЫ И СПРАВОЧНИК
    # -------------------------------------------------------------
    ws_set = wb.create_sheet(title="Регламенты и Спецификации")
    ws_set.views.sheetView[0].showGridLines = True
    
    ws_set["B2"] = f"ТЕХНИЧЕСКИЙ РЕГЛАМЕНТ И ИНТЕРВАЛЫ: {vehicle.get('name', '')}"
    ws_set["B2"].font = font_title
    ws_set["B3"] = f"Спецификация масла: {vehicle.get('oil_spec', '-')} | VIN: {vehicle.get('vin', '-')}"
    ws_set["B3"].font = font_subtitle
    
    set_headers = ["№", "ID", "Расходник / Узел", "Категория", "Интервал (км)", "Интервал (м/ч)", "Предупреждение (км)", "Спецификация / Артикул"]
    for ci, h in enumerate(set_headers, start=2):
        cell = ws_set.cell(row=5, column=ci, value=h)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_bottom_navy
        
    for r_idx, t in enumerate(trackers, start=6):
        row_vals = [
            r_idx - 5,
            t.get("id", ""),
            t.get("name", ""),
            t.get("category", ""),
            t.get("interval_km", 0),
            t.get("interval_hours", 0) if t.get("interval_hours", 0) > 0 else "-",
            t.get("warn_km", 1500),
            t.get("article") or t.get("spec") or "-"
        ]
        for ci, val in enumerate(row_vals, start=2):
            cell = ws_set.cell(row=r_idx, column=ci, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if ci in [6, 8]:
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0 "км"'
            elif ci in [2, 3, 5, 7]:
                cell.alignment = Alignment(horizontal="center")
                
    # Auto-adjust column widths for all sheets
    for ws in [ws_charts, ws_status, ws_log, ws_set]:
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col[0].column == 1: # Column A margin
                ws.column_dimensions[col_letter].width = 3
                continue
            max_len = 0
            for cell in col:
                if cell.row in [2, 3]: continue # ignore big header text
                val_str = str(cell.value or '')
                if not val_str.startswith('='):
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_file.close()
    wb.save(temp_file.name)
    
    clean_brand = "".join(c for c in vehicle.get('brand', 'авто') if c.isalnum() or c in (' ', '_', '-')).strip()
    clean_model = "".join(c for c in vehicle.get('model', '') if c.isalnum() or c in (' ', '_', '-')).strip()
    filename = f"Обслуживание_{clean_brand}_{clean_model}.xlsx"
    
    # URL encoded filename for RFC 5987 compatibility in Content-Disposition
    encoded_filename = urllib.parse.quote(filename)
    
    return FileResponse(
        temp_file.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

# --- BACKUP / SYNC ENDPOINTS (100% COMPATIBLE WITH MOBILE APP) ---
@app.get("/api/export-json")
@app.get("/api/backup/export")
def export_json():
    db = load_db()
    vehicle = get_active_vehicle(db)
    vehicles = db.setdefault("vehicles", [])
    
    # Ensure active vehicle is up to date in vehicles array
    for v in vehicles:
        if v.get("id") == vehicle.get("id"):
            v.update(vehicle)
            
    backup_data = {
        "version": "2.5",
        "app": "car-maintenance-app",
        "exported_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "active_vehicle_id": db.get("active_vehicle_id", "car_1"),
        "vehicle": vehicle,
        "vehicles": vehicles,
        "trackers": get_vehicle_trackers(db, vehicle),
        "maintenance_records": db.get("maintenance_records", []),
        "reference_intervals": db.get("reference_intervals", [])
    }
    
    clean_brand = "".join(c for c in vehicle.get('brand', 'авто') if c.isalnum() or c in (' ', '_', '-')).strip()
    clean_model = "".join(c for c in vehicle.get('model', '') if c.isalnum() or c in (' ', '_', '-')).strip()
    filename = f"backup_{clean_brand}_{clean_model}.json"
    encoded_filename = urllib.parse.quote(filename)
    
    return JSONResponse(
        content=backup_data,
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

@app.post("/api/import-json")
@app.post("/api/backup/import")
def import_json(payload: Dict[str, Any], auth: bool = Depends(require_admin)):
    if not payload:
        raise HTTPException(status_code=400, detail="Пустой файл бэкапа")
        
    db = load_db()
    
    # Restore vehicles
    if "vehicles" in payload and isinstance(payload["vehicles"], list) and len(payload["vehicles"]) > 0:
        db["vehicles"] = payload["vehicles"]
    elif "vehicle" in payload and isinstance(payload["vehicle"], dict):
        v = payload["vehicle"]
        if "id" not in v:
            v["id"] = "car_1"
        db["vehicles"] = [v]
        
    if "active_vehicle_id" in payload:
        db["active_vehicle_id"] = payload["active_vehicle_id"]
    elif "vehicles" in db and len(db["vehicles"]) > 0:
        db["active_vehicle_id"] = db["vehicles"][0].get("id", "car_1")
        
    # Restore vehicle main object
    if "vehicle" in payload and isinstance(payload["vehicle"], dict):
        db["vehicle"] = payload["vehicle"]
    elif "vehicles" in db and len(db["vehicles"]) > 0:
        db["vehicle"] = db["vehicles"][0]
        
    # Restore maintenance records
    if "maintenance_records" in payload and isinstance(payload["maintenance_records"], list):
        # Normalize records: ensure ids and vehicle_ids
        records = []
        for idx, r in enumerate(payload["maintenance_records"], start=1):
            if not isinstance(r, dict): continue
            r_norm = dict(r)
            if "id" not in r_norm or not r_norm["id"]:
                r_norm["id"] = idx
            if "vehicle_id" not in r_norm or not r_norm["vehicle_id"]:
                r_norm["vehicle_id"] = db.get("active_vehicle_id", "car_1")
            records.append(r_norm)
        db["maintenance_records"] = records
        
    # Restore trackers
    if "trackers" in payload and isinstance(payload["trackers"], list):
        db["trackers"] = payload["trackers"]
        
    if "reference_intervals" in payload and isinstance(payload["reference_intervals"], list):
        db["reference_intervals"] = payload["reference_intervals"]
        
    save_db(db)
    
    return {
        "status": "success",
        "message": "База данных и настройки успешно восстановлены!",
        "vehicles_count": len(db.get("vehicles", [])),
        "records_count": len(db.get("maintenance_records", [])),
        "active_vehicle_id": db.get("active_vehicle_id")
    }


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8080))
    host = os.environ.get("HOST", "0.0.0.0")
    print(f"🚀 Starting Auto Maintenance Server on http://localhost:{port}")
    uvicorn.run("app:app", host=host, port=port, reload=False)
